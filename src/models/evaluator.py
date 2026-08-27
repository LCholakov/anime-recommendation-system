import pandas as pd
from typing import Callable
from datetime import datetime


def evaluate(
    recommend_fn: Callable,
    test_df: pd.DataFrame,
    train_df: pd.DataFrame,
    n: int = 10
) -> dict:
    precisions, recalls, hits = [], [], []

    for user_id, test_group in test_df.groupby("user_id"):
        recs    = recommend_fn(user_id, n=n)
        rec_ids = set(recs["anime_id"])
        test_ids = set(test_group["anime_id"])

        hit       = len(rec_ids & test_ids) > 0
        precision = len(rec_ids & test_ids) / n
        recall    = len(rec_ids & test_ids) / len(test_ids) if test_ids else 0.0

        hits.append(hit)
        precisions.append(precision)
        recalls.append(recall)

    if not hits:
        return {"hit_rate": 0.0, "precision": 0.0, "recall": 0.0}
    return {
        "hit_rate":  round(sum(hits) / len(hits), 4),
        "precision": round(sum(precisions) / len(precisions), 4),
        "recall":    round(sum(recalls) / len(recalls), 4),
    }


def log_run(
    xlsx_path: str,
    model_name: str,
    metrics: dict,
    hyperparams: dict,
    comment: str,
) -> None:
    """Append one run entry to the xlsx performance tracker.

    Each call adds a new row — intermediate and failed runs are recorded too.
    The row includes an auto-incremented Run #, timestamp, model name,
    all hyperparams (epochs, batch_size, n_components, etc.), metrics, and comment.

    Args:
        xlsx_path:   path to report/model_performance_tracker.xlsx
        model_name:  e.g. "SVD Collaborative Filtering"
        metrics:     dict with hit_rate, precision, recall (from evaluate())
        hyperparams: dict — keys match tracker columns (split, min_ratings,
                     n_recommendations, epochs, batch_size, patience, embed_dim,
                     n_components, train_users). Missing keys default to "N/A".
        comment:     free-text note about what changed / why this run was done
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    NA = "N/A"
    row_data = [
        None,           # Run # — filled below after reading current max
        timestamp,
        model_name,
        hyperparams.get("split",                 "leave-one-out"),
        hyperparams.get("min_ratings",           5),
        hyperparams.get("n_recommendations",     10),
        hyperparams.get("m_percentile",          NA),
        hyperparams.get("min_rating_threshold",  NA),
        str(hyperparams["sublinear_tf"]) if "sublinear_tf" in hyperparams else NA,
        hyperparams.get("epochs",                NA),
        hyperparams.get("batch_size",            NA),
        hyperparams.get("patience",              NA),
        hyperparams.get("embed_dim",             NA),
        hyperparams.get("n_components",          NA),
        hyperparams.get("train_users",           NA),
        hyperparams.get("train_secs",            NA),
        metrics.get("hit_rate",  0.0),
        metrics.get("precision", 0.0),
        metrics.get("recall",    0.0),
        comment,
    ]
    headers = [
        "Run #", "Timestamp", "Model", "Split",
        "min_ratings", "n_recommendations",
        "m_percentile", "min_rating_threshold", "sublinear_tf",
        "epochs", "batch_size", "patience", "embed_dim", "n_components", "train_users",
        "Train time (s)",
        "Hit Rate @10", "Precision @10", "Recall @10",
        "What changed / Comment",
    ]
    # read current max run number before appending
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    max_run = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, (int, float)):
            max_run = max(max_run, int(val))
    row_data[0] = max_run + 1
    wb.close()

    append_to_tracker(xlsx_path, row_data, headers)


def append_to_tracker(xlsx_path: str, row_data: list, headers: list) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_font   = Font(bold=True, color="FFFFFF")
    header_fill   = PatternFill("solid", fgColor="2F5496")
    row_fill      = PatternFill("solid", fgColor="E2EFDA")
    border_side   = Side(style="thin", color="CCCCCC")
    cell_border   = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet_is_empty = ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None

    if sheet_is_empty:
        # Write full header row from scratch
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = cell_border
        ws.row_dimensions[1].height = 30
        col_index = {h: i + 1 for i, h in enumerate(headers)}
    else:
        # Read existing header row to build column-name → column-number map.
        # Add any new headers that don't exist yet (e.g. "Train time (s)").
        col_index = {}
        for cell in ws[1]:
            if cell.value is not None:
                col_index[cell.value] = cell.column
        next_col = max(col_index.values()) + 1 if col_index else 1
        for h in headers:
            if h not in col_index:
                cell = ws.cell(row=1, column=next_col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = cell_border
                col_index[h] = next_col
                next_col += 1

    next_row = ws.max_row + 1 if ws.cell(1, 1).value is not None else 2

    # Write each value into the correct column by name
    for h, val in zip(headers, row_data):
        col = col_index.get(h)
        if col is None:
            continue
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.fill = row_fill
        cell.alignment = center
        cell.border = cell_border

    # Auto-fit column widths across all used columns
    for h, col in col_index.items():
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, next_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    wb.save(xlsx_path)

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.data_work.data_loader import load_anime_data, load_rating_data
from src.models.baseline import split_train_test, compute_bayesian_scores, recommend_popular_anime, evaluate_model

# --- load & split ---
ratings_df = load_rating_data("data/rating_clean.csv")
anime_df   = load_anime_data("data/anime_clean.csv")

print("Splitting train/test (min 5 ratings per user, 70/30)...")
train, test = split_train_test(ratings_df, min_ratings=5, test_ratio=0.3)
print(f"  Train: {len(train)} rows | Test: {len(test)} rows")

# --- compute scores ---
print("Computing Bayesian scores from training data...")
scores = compute_bayesian_scores(train)
scores_named = scores.merge(anime_df[["anime_id", "name", "genre"]], on="anime_id", how="left")
scores_named = scores_named.sort_values("bayesian_score", ascending=False).reset_index(drop=True)

# --- evaluate ---
print("Evaluating (n=10)...")
metrics = evaluate_model(scores, test, train, n=10)
print(f"  Hit Rate : {metrics['hit_rate']}")
print(f"  Precision: {metrics['precision']}")
print(f"  Recall   : {metrics['recall']}")

# --- save data ---
train.to_csv("data/train.csv", index=False)
test.to_csv("data/test.csv", index=False)
scores_named.to_csv("data/baseline_scores.csv", index=False)
print(f"✅ Saved data/train.csv, data/test.csv, data/baseline_scores.csv")

print("\nTop 10 by Bayesian score:")
print(scores_named[["name", "rating_count", "avg_rating", "bayesian_score"]].head(10).to_string(index=False))

# --- write to model performance tracker ---
xlsx_path = "report/model_performance_tracker.xlsx"
wb = load_workbook(xlsx_path)
ws = wb.active

# styles
header_font    = Font(bold=True, color="FFFFFF")
header_fill    = PatternFill("solid", fgColor="2F5496")
baseline_fill  = PatternFill("solid", fgColor="D9E1F2")
border_side    = Side(style="thin", color="CCCCCC")
cell_border    = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
center         = Alignment(horizontal="center", vertical="center", wrap_text=True)

headers = [
    "Model", "min_ratings", "test_ratio", "n_recommendations",
    "m (percentile)", "Hit Rate @10", "Precision @10", "Recall @10", "Comments"
]

# write headers if sheet is empty
if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = cell_border
    ws.row_dimensions[1].height = 30

# find next empty row
next_row = ws.max_row + 1
if ws.max_row == 1 and ws.cell(1, 1).value is None:
    next_row = 2

row_data = [
    "Baseline (Popularity)",
    5,          # min_ratings
    0.3,        # test_ratio
    10,         # n_recommendations
    "80th pct", # m threshold
    metrics["hit_rate"],
    metrics["precision"],
    metrics["recall"],
    "Greedy baseline. Recommends same top-10 popular anime to all users. No personalisation."
]

for col, val in enumerate(row_data, 1):
    cell = ws.cell(row=next_row, column=col, value=val)
    cell.fill = baseline_fill
    cell.alignment = center
    cell.border = cell_border

# auto-width columns
for col in range(1, len(headers) + 1):
    max_len = max(
        len(str(ws.cell(row=r, column=col).value or ""))
        for r in range(1, next_row + 1)
    )
    ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

wb.save(xlsx_path)
print(f"✅ Results written to {xlsx_path}")
